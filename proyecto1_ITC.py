# Aqui se irá actualizando el proyecto para su total funcionalidad a lo largo del curso de ITC

"""
ENTRADAS:
    1. Solicitar al usuario el formato excel preestablecido con el nombre/clave de los productos, costo y precio.

PROCESO:
    1. Ingresar el Excel_productos
    2. Utilizar pandas para agregar una tercera columna después de (nombre/clave, precio_venta, costo)
    3. En la cuarta columna creada por pandas establecer una fórmula de Margen = (precio_venta - costo)/precio_venta
        3.1 Formato en porcentaje para la columna 4
        3.2 Usar filtro para organizar las columnas por margen de mayor a menor
    4. Crear una quinta columna con formula promedio_margen = (SUM(Margen)/COUNT(Margen))
    5. Importar informacion como CSV a streamlit
    6. Crear variable diferencia = ((margen - promedio margen)/promedio margen)*100
    7. Crear visualizacion de gráfico de barras con reglas:
        7.1 While diferencia >= 5 pintar en verde la columna del producto por que tiene margen sobresaliente
        7.2 While diferencia <= -5 pintar en rojo la columna del producto por que tiene margen subsaliente
        7.3 SINO pintar en azul la columna del producto por que el margen esta dentro del rango

SALIDAS: 
    Gráfico de barras dinámico que muestre tus porductos junto con una linea del margen promedio y cambia el color de los productos sobre o sub 
    salientes dentro de tu inventario ordenados de mayor a menor.
        
"""

import tkinter as tk #Documentation: https://docs.python.org/3/library/tk.html
from tkinter import filedialog
import pandas as pd#Documentation: https://pandas.pydata.org/docs/
import streamlit #Documentation: https://docs.streamlit.io/
import openpyxl #Documentation: https://openpyxl.readthedocs.io/en/stable/ 
from openpyxl import load_workbook
from openpyxl import workbook

root = tk.Tk()
root.withdraw()

file = filedialog.askopenfilename( ### pops up the file selector window from tkinter
    title="Select a file", ### the name that gives the user the instruction.
    filetypes=[("Archivos de Excel", "*.xlsx * .xls")] ### discriminates the files that are slectable 
)

# We validate that the file is an excel file

if file.endswith((".xlsx", ".xls")): ### only continues the programm if the file is in the correct type
    try:
        data = pd.read_excel(file) ### data to pandas to forward analisys
        workbook = load_workbook(file) ### variable saved with the excel file selected
        work_sheet = workbook.active ### gets the active sheet inside the excel file
        for line in range(2,work_sheet.max_row + 1): ### Writes the formula below in every cell that has the info requiered
            margin = f"=(B{line}-C{line})/B{line}" #### is the formula that is going to be inserted in every cell until the last cell with data in order to get the margin
            work_sheet[f"D{line}"] = margin ### adds the formula into the 3 column 
            work_sheet[f"D{line}"].number_format = '0.00%' ### turns the float number into percetage format
        work_sheet["D1"] = "Margen" ### names the first cell of the column as "Margen"
        workbook.save("resultadoprueba1.xlsx") ### saves the modified document under the name established
    except Exception as error:
        print(f"An error has occurred :{error}") ### in case file has a problem with reading prints the error to the user
else:
    print("The file is not compatible") ### in case the file selected is not from the specified type.


        
